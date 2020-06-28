from openpyxl import load_workbook
from .transaction import Transaction
VALUE = 3
DATE = 0
DESC = 6
SENDER = 5
TYPE = 8

decode_dict = {"\u0105": "a",
               "\u0107": "c",
               "\u0119": "e",
               "\u0142": "l",
               "\u0144": "n",
               "\u01F3": "o",
               "\u015B": "s",
               "\u017A": "z",
               "\u017C": "z",
               "\u0104": "A",
               "\u0106": "C",
               "\u0118": "E",
               "\u0141": "L",
               "\u0143": "N",
               "\u00D3": "O",
               "\u015A": "S",
               "\u0179": "Z",
               "\u017B": "Z"}


def desc_decode(desc: str):
    if(isinstance(desc, str)):
        for key, value in decode_dict.items():
            desc = desc.replace(key, value)
    return desc


class Parser:
    def __init__(self):
        pass

    @classmethod
    def _get_target(cls, row):
        type_of_trans = desc_decode(row[TYPE].value)
        target = "unknown"
        if(type_of_trans in ['Przelew wychodzacy', 'Przelew przychodzacy', 'Zlecenie stale']):
            target = row[SENDER].value.splitlines()[1]
        elif(type_of_trans in ['Transakcja karta']):
            target = " ".join(row[DESC].value.split()[3: -4])
        elif(type_of_trans in ['Transakcja BLIK']):
            target = row[DESC].value.split(",")[2]

        return desc_decode(target.strip())

    @classmethod
    def _parse_row(cls, row):
        target = cls._get_target(row)
        value = float(row[VALUE].value)
        date = row[DATE].value.strftime("%Y-%m-%d")
        if value < 0:
            trans = Transaction(date, -value, target)
        else:
            trans = Transaction(date, value, target, 100)

        return trans

    @classmethod
    def getTransaction(cls, xls_file):
        xls = load_workbook(xls_file).active
        sheetX = xls["A:L"]
        trans_list = []
        for row in xls.iter_rows(min_row=2, max_col=12):
            trans_list.append(cls._parse_row(row))
        #sheetX.apply(cls._parse_row, axis=1, trans_list=trans_list)
        return trans_list
